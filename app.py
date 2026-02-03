from flask import Flask, render_template, request, send_file
import pandas as pd
from openpyxl import load_workbook
import tempfile
import zipfile
import os

app = Flask(__name__)

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/process", methods=["POST"])
def process():
    csv_file = request.files["csv_file"]
    excel_template = request.files["excel_template"]

    df = pd.read_csv(csv_file)

    # Tạo file ZIP tạm
    zip_path = tempfile.NamedTemporaryFile(delete=False, suffix=".zip").name

    with zipfile.ZipFile(zip_path, "w") as zipf:
        for i in range(len(df)):
            row = df.iloc[i]

            # Tạo file Excel tạm cho từng dòng
            temp_excel = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsm").name

            # Mở template
            wb = load_workbook(excel_template, keep_vba=True)
            ws = wb.active

            # Điền dữ liệu text
            ws["B3"] = row["Date demande"]
            ws["G3"] = row["Date mise en service"]
            ws["G5"] = row["Nb expéditions"]
            ws["G6"] = row["Nb colis"]
            ws["B7"] = row["Produit"]
            ws["B9"] = row["Agence"]
            ws["B14"] = row["Commercial"]
            ws["G14"] = row["Présence installation"]
            ws["B16"] = row["Raison sociale"]
            ws["G16"] = row["SIRET administratif"]
            ws["B18"] = row["Adresse d'installation"]
            ws["B21"] = row["UTILISATEUR"]
            ws["F21"] = row["DECIDEUR"]
            ws["B22"] = row["Téléphone UTILISATEUR"]
            ws["F22"] = row["Téléphone DECIDEUR"]
            ws["B23"] = row["Email"]
            ws["A26"] = row["Compte ERM"]
            ws["B26"] = row["Nom"]

            # Lưu file Excel tạm
            wb.save(temp_excel)

            # Đặt tên file theo số dòng hoặc theo dữ liệu
            filename = f"Demande_{i+1}.xlsm"

            # Thêm file Excel vào ZIP
            zipf.write(temp_excel, arcname=filename)

            # Xóa file Excel tạm
            os.remove(temp_excel)

    # Trả file ZIP về cho người dùng
    return send_file(zip_path, as_attachment=True, download_name="resultats.zip")

if __name__ == "__main__":
    app.run()
